"""01.07.2024, Version 1.0.

Creates Relevant images for certain turbine.

@author: Stefan Janisch

"""

# python modules
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import concurrent.futures
import multiprocessing
import time
import sqlite3
import sys
project_dir = os.path.dirname(os.path.dirname(__file__))
# Insert project directory to path
if project_dir not in sys.path:
    sys.path.insert(0, project_dir)
from options import *

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule

from data_cleaning.compute_power_and_flag import (get_turbine_data_from_id,
                                    get_turbine_power_curve_from_type, print_green)

path = FolderPaths.data_dir

def mask_diagnostics(turbine: str, path: str, threshold: float, mode: str = "quantiles", saveplot: bool = True) -> tuple[dict, dict, int, bool, str, str]:
    """Runs diagnostics for the flagging of the turbine specified via "turbine".
    Counts the number of flags per turbine. If the number of flags exceeds a certain threshold,
    the turbine is flagged as suspicious.

    Variables
    ---------
    turbine : str
        Turbine ID
    path : str
        Path to data folder
    threshold : float
        Threshold for flagging a turbine as suspicious. If the ratio of flagged data points
        to total data points is above this threshold, the turbine is flagged as suspicious.
    mode : str
        Mode for flagging. Either "quantiles" or "goretti".
    saveplot : bool
        Whether to save the plot or not.
    
    Returns
    -------
    counts_flagged : dict
        Dictionary containing the number of flagged data points per flag.
    ratios : dict
        Dictionary containing the ratio of flagged data points to total data points per flag.
    datapoints_num : int
        Total number of data points for the turbine.
    sus_turbine : bool
        Whether the turbine is suspicious or not.
    filepath : str 
        Path to the plot.
    turbine_data[0] : str
        The turbine type.
    """
    
    print(f"Running mask diagnostics for turbine {turbine}")

    # read in data
    data = pd.read_csv(os.path.join(FolderPaths.flagged_dir, turbine + ".csv"), encoding="utf-8")
    
    windspeed_data = data[data["Flag Computed"] <= 1]  # extract data for wind speed distribution
    data = data.dropna(subset=["windspeed [m/s]", "power [kW]"]) # drop NaN values

    # get powercurve
    turbine_data = get_turbine_data_from_id(path=path, turbine_id=turbine)
    powercurve_df = get_turbine_power_curve_from_type(path=path, turbine_type=turbine_data[0])
    powercurve_df["value"] = powercurve_df["value"] * 1e-3

    # Get nominal power
    nominal_power = turbine_data[1]

    # read cleaning_config file
    dc_config_path = os.path.join(FolderPaths.metadata_dir, "data_cleaning_config.xlsx")
    dc_config = pd.read_excel(dc_config_path)
    # get cleaning config for turbine type
    turbine_config = dc_config[dc_config["turbine_type"] == turbine_data[0]]
    # Read sc_mult
    sc_mult = turbine_config["sc_mult"].values[0]

    # count occurences of different flags and write them into a dictionary
    counts_flagged = data["Flag Computed"].value_counts().to_dict()
    # count total number of data points
    datapoints_num = len(data)
    # make dictionary of ratios btw flagged and total data points
    ratios = {key: counts_flagged[key]/datapoints_num for key in counts_flagged.keys()}
    
    # check for suspicious ratios
    if ratios[0] < (1 - threshold/100):
        sus_turbine = True
    else:
        sus_turbine = False

    # group data by flag
    grouped_data = data.groupby("Flag Computed")  

    if saveplot:
        # Add histogram of flagged data
        f, (ax_power, ax_hist) = plt.subplots(1, 2, sharey=True, gridspec_kw={'width_ratios': [4, 1]}, figsize=(16,9))

        # Define colors for each flag
        colors = {
            0: 'C0',
            1: 'C1',
            2: 'C2',
            3: 'C5',
            4: 'C6',
            5: 'C3',
            6: 'C4'
        }

        # Create stacked histogram
        ax_hist.hist([group["power [kW]"] for name, group in grouped_data if name != 0], 
                        bins=round(nominal_power/20),
                        density=True,
                        orientation="horizontal",
                        color=[colors[name] for name, _ in grouped_data if name != 0],
                        stacked=True)

        ax_hist.set_xlabel("log density")
        ax_hist.set_title("Distribution of flagged data points")
        ax_hist.set_xscale("log")

        # Add power curves with flag
        for name, group in grouped_data:
            if name == 0:
                labelname = "valid data"
            elif name == 1:
                labelname = "value out of sensible range"
            elif name == 2 and mode == "quantiles":
                labelname = "wind speed out of quantile (with not-negative derivative)"
            elif name == 2 and mode == "goretti":
                labelname = "wind speed higher than cut out - power above zero \nwind speed lower than cut in - power above 4 %"
            elif name == 3 and mode == "goretti":
                labelname = "wind speed between cut in and rated - power below 4 % \nwind speed between rated and cut out - power below 95 %"
            elif name == 4 and mode == "goretti":
                labelname = "power out of 3rd-quantile of wind speed bin."
            if name == 5:
                labelname = f"power outside {sc_mult} sigma range"
            if name == 6:
                labelname = "flagged by power curve shifting"
            labelname = labelname + str(f" ({round(ratios[name]*100)}%)")
            ax_power.plot(group["windspeed [m/s]"], group["power [kW]"], marker=".", linestyle="", markersize=1, label=labelname, alpha=0.6, color=colors[name])
        ax_power.plot(powercurve_df["wind_speed"], powercurve_df["value"], color="k", label="power curve from provider", alpha=0.3, linestyle="dashed")
        ax_power.set_xlabel("wind speed [m/s]")
        ax_power.set_ylabel("power [kW]")
        pos = ax_power.get_position()
        ax_power.set_position([pos.x0, pos.y0, pos.width, pos.height * 0.7])
        ax_power.legend(markerscale = 10, fontsize="small", loc="lower right")
        ax_power.set_title(f"Flagged values for Turbine ID {turbine}: {round((1-ratios[0])*100, 2)}% flagged")
        f.subplots_adjust(wspace=0)
        plt.tight_layout

        # Enable grid for both plots
        ax_power.grid()
        ax_hist.grid()

        filepath = os.path.join(FolderPaths.diagnostics_dir, "FlagEv-" + turbine + ".png")
        plt.savefig(filepath, dpi=400)
        if sus_turbine:
            plt.savefig(os.path.join(FolderPaths.diagnostics_dir, "SuspiciousTurbines", "FlagEv-" + turbine + ".png"), dpi=400)
        plt.close()

    # extract only wind speed data and exclude nan
    windspeeds = windspeed_data["windspeed [m/s]"]
    windspeeds = windspeeds[~windspeeds.isna()]

    #print(f"Should be flagged: {should_be_flagged_num}")
    #print(f"Wrongly flagged: {wrongly_flagged_num}")
    return counts_flagged, ratios, datapoints_num, sus_turbine, filepath, turbine_data[0]


# Define a function to process a single file
def process_file(file: str) -> tuple[str, list[int], list[float], int, bool, str, str]:
    """Wrapper for file processing. Generates path for turbine file and
    performs mask_diagnostics on the file.
    """    

    this_file = os.path.join(path, "TimeSeriesPrepared", file)
    turbine_id = os.path.splitext(os.path.basename(this_file))[0]
    counts_flagged, ratios, datapoints_num, sus_turbine, filepath, turbine_type = mask_diagnostics(turbine_id, path, Diagnostics.threshold_flagging, saveplot=True)
    return turbine_id, counts_flagged, ratios, datapoints_num, sus_turbine, filepath, turbine_type

def fetch_turbine_type(turbine_id: str, database_path: str) -> str:
    """Returns the turbine type for the given turbine_id from the database at database_path."""
    # Get turbine type for turbine ID
    verbindung = sqlite3.connect(database_path)
    zeiger = verbindung.cursor()
    zeiger.execute("SELECT * FROM metadaten WHERE \"turbine ID\" = \"" + turbine_id + "\"")
    turbine_metadata = zeiger.fetchall()

    return turbine_metadata[0][3]

# Define the main function
def main():
    """Main function. Iterates through all files in TimeSeriesPrepared
    and performs mask diagnostics across all CPU cores of the host
    system.
    """    

    turbine_files = os.listdir(FolderPaths.flagged_dir)

    # Check if output folder exists. If not, create it
    if not os.path.exists(FolderPaths.diagnostics_dir):
        os.mkdir(FolderPaths.diagnostics_dir)
    if not os.path.exists(os.path.join(FolderPaths.diagnostics_dir, "SuspiciousTurbines")):
        os.mkdir(os.path.join(FolderPaths.diagnostics_dir, "SuspiciousTurbines"))

    # Create a ProcessPoolExecutor for concurrent execution
    print_green(f"Running mask diagnostics in a pool of {ParallelProcessing.num_conc_proc} processes")
    print("This may take a while ..................................")
    with concurrent.futures.ProcessPoolExecutor(max_workers=ParallelProcessing.num_conc_proc) as executor:
        # Process files using separate processes and collect results
        results = list(executor.map(process_file, turbine_files))
        

    # Unpack the results into separate lists
    list_turbine_id, list_counts_flagged, list_ratios, list_datapoints_num, list_sus_turbine, list_filepath, list_turbine_type = [], [], [], [], [], [], []
    for turbine_id, counts_flagged, ratios, datapoints_num, sus_turbine, filepath, turbine_type in results:
        list_turbine_id.append(turbine_id)
        list_counts_flagged.append(counts_flagged)
        list_ratios.append(ratios)
        list_datapoints_num.append(datapoints_num)
        list_sus_turbine.append(sus_turbine)
        list_filepath.append(filepath)
        list_turbine_type.append(turbine_type)

    # WRITE RESULTS TO EXCEL FILE #############################################
    print_green("Writing to Excel")
    excel_path = os.path.join(FolderPaths.diagnostics_dir, "Flag Evaluation Results.xlsx")

    # Prepare a pandas DataFrame

    results_dict = {'Turb. ID': list_turbine_id, 'Turb. type': list_turbine_type, 'Sus. Turb.': list_sus_turbine, 'File Link': list_filepath, 'Total dp.': list_datapoints_num}
    for key in [0,1,2,5,6]:
        results_dict[f"Ct. ({key})"] = [dic[key] if key in dic.keys() else 0 for dic in list_counts_flagged ]
        results_dict[f"Rt. ({key})"] = [dic[key] if key in dic.keys() else 0 for dic in list_ratios]

    Diag_results = pd.DataFrame(results_dict)

    # Rename columns
    Diag_results.rename(columns={'Ct. (0)': 'Ct. (valid)', 'Rt. (0)': 'Rt. (valid)',
                                 'Ct. (1)': 'Ct. (rng)', 'Rt. (1)': 'Rt. (rng)',
                                 'Ct. (2)': 'Ct. (quant)', 'Rt. (2)': 'Rt. (quant)',
                                 'Ct. (5)': 'Ct. (sc)', 'Rt. (5)': 'Rt. (sc)', 
                                 'Ct. (6)': 'Ct. (pc)', 'Rt. (6)': 'Rt. (pc)'}, inplace=True)

    # Write Data Frame to Excel
    Diag_results.to_excel(excel_path)
    
    # EXCEL FORMATTING ########################################################

    # Link plots to workbook
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.active

    # Freeze first row
    worksheet.freeze_panes = "A2"

    # Set filters for columns A to P
    worksheet.auto_filter.ref = f"A1:P{worksheet.max_row}"

    # Set column widths to 11 for all columns
    for col in worksheet.columns:
        worksheet.column_dimensions[col[0].column_letter].width = 11
    
    # First row left-aligned and bold
    for cell in worksheet[1]:
        cell.alignment = openpyxl.styles.Alignment(horizontal="left")
        cell.font = openpyxl.styles.Font(bold=True)


    for cell in worksheet['E']:
        if cell.value != "File Link":
            linkpath = str(cell.value)
            cell.value = '=HYPERLINK("{}", "{}")'.format("file:///"+linkpath, "Plot")
            # Format in blue
            cell.font = openpyxl.styles.Font(underline="single", color="0000FF")

    # Color Suspicious Turbines red
    for cell in worksheet['D']:
        if cell.value == True:
            cell.font = openpyxl.styles.Font(color="00ff0000")
            cell.offset(column=-2).font = openpyxl.styles.Font(color="00ff0000")
            cell.offset(column=-1).font = openpyxl.styles.Font(color="00ff0000")
    
    # Conditional formatting for ratios
    for col in ["H"]:
        for cell in worksheet[col][1:]:
            if isinstance(cell.value, str) and '%' in cell.value:
                cell.value = float(cell.value.strip("%"))
            elif isinstance(cell.value, (int, float)):
                cell.value = cell.value
            else:
                # Handle the case where cell.value is neither a string with '%' nor a number
                continue

            cell.number_format = "0.00%"

        # Define the range for conditional formatting for the entire column (excluding the header)
        column_range = f"{col}2:{col}{worksheet.max_row}"

        # Create a 3-color scale rule
        rule1 = ColorScaleRule(start_type='min', start_color='FF0000', 
                            mid_type='percentile', mid_value=50, mid_color='FFFF00', 
                            end_type='max', end_color='00FF00')

        # Add the rule to the worksheet
        worksheet.conditional_formatting.add(column_range, rule1)

    for col in ["J", "L", "N", "P"]:
        for cell in worksheet[col][1:]:
            if isinstance(cell.value, str) and '%' in cell.value:
                cell.value = float(cell.value.strip("%"))
            elif isinstance(cell.value, (int, float)):
                cell.value = cell.value
            else:
                # Handle the case where cell.value is neither a string with '%' nor a number
                continue

            cell.number_format = "0.00%"

        # Define the range for conditional formatting for the entire column (excluding the header)
        column_range = f"{col}2:{col}{worksheet.max_row}"

        # Create a 3-color scale rule
        rule2 = ColorScaleRule(start_type='min', start_color='00FF00', 
                            mid_type='percentile', mid_value=50, mid_color='FFFF00', 
                            end_type='max', end_color='FF0000')

        # Add the rule to the worksheet
        worksheet.conditional_formatting.add(column_range, rule2)
    
    # Output threshold 0 in cell R2
    worksheet['R1'].value = "Threshold for susp. turbine:"
    worksheet['R1'].font = openpyxl.styles.Font(bold=True)
    worksheet['R2'].value = Diagnostics.threshold_flagging / 100
    worksheet['R2'].number_format = "0.00%"


    # Save workbook
    workbook.save(excel_path)
    workbook.close()

